# utils.py - ИСПРАВЛЕННАЯ цветовая маркировка (читает конкретные строки)

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import pickle
from typing import Any
import os
import shutil

def normalize_text(text) -> str:
    """Нормализация текста для поиска"""
    if not isinstance(text, str):
        return text
    text = text.lower()
    text = re.sub(r'[^a-zа-я0-9\s]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def clean_product_name_advanced(raw_text: str) -> str:
    """УЛУЧШЕННАЯ очистка названия товара от доп. параметров"""
    if not raw_text or not isinstance(raw_text, str):
        return ""
    
    text = raw_text.strip()
    
    remove_patterns = [
        r'возможность\s+поставки\s+аналогов\s*:\s*\w+',
        r'валюта\s*:\s*\w+',
        r'единица\s+измерения\s*:\s*\w+',
        r'страна\s+происхождения\s*:\s*[^\n]+',
        r'производитель\s*:\s*[^\n]+',
        r'гарантия\s*:\s*[^\n]+',
        r'срок\s+поставки\s*:\s*[^\n]+',
        r'количество\s*:\s*\d+',
        r'цена\s*:\s*[^\n]+',
        r'артикул\s*:\s*[^\n]+',
        r'код\s+товара\s*:\s*[^\n]+',
    ]
    
    for pattern in remove_patterns:
        text = re.sub(pattern, '', text, flags=re.IGNORECASE)
    
    text = re.sub(r'\n+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    text = text.strip()
    
    exclude_lines = [
        r'^возможность', r'^валюта', r'^единица', r'^страна',
        r'^производитель', r'^гарантия', r'^срок', r'^количество',
        r'^цена', r'^артикул', r'^код\s+товара', r'^\d+\s*$', r'^[a-z]{2,3}\s*$',
    ]
    
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    clean_lines = []
    
    for line in lines:
        is_excluded = False
        for exclude_pattern in exclude_lines:
            if re.match(exclude_pattern, line, re.IGNORECASE):
                is_excluded = True
                break
        if not is_excluded and len(line) > 3:
            clean_lines.append(line)
    
    if clean_lines:
        result = clean_lines[0]
        result = re.sub(r'[^\w\s,.-]', '', result)
        result = re.sub(r'\s+', ' ', result).strip()
        return result
    
    return text

def parse_price_value(price_str: str) -> float:
    """Извлекает числовое значение цены из строки"""
    if not price_str or not isinstance(price_str, str):
        return 0.0
    
    try:
        # Убираем все кроме цифр, запятых и точек
        clean_price = re.sub(r'[^\d,.]', '', price_str)
        clean_price = clean_price.replace(',', '.')
        clean_price = clean_price.replace(' ', '')
        
        # Убираем множественные точки
        if clean_price.count('.') > 1:
            parts = clean_price.split('.')
            clean_price = ''.join(parts[:-1]) + '.' + parts[-1]
        
        return float(clean_price) if clean_price else 0.0
    except:
        return 0.0

def get_color_by_comparison(yandex_price: float, min_tender_price: float) -> str:
    """
    Определяет цвет ячейки по сравнению с минимальной тендерной ценой
    
    - Зеленый: нет превышения (цена ЯМ <= минимальной тендерной + 5%)
    - Желтый: превышение 5-10%
    - Красный: превышение > 10%
    """
    if min_tender_price == 0 or min_tender_price == float('inf'):
        return "00FF00"  # Зеленый по умолчанию если нет данных
    
    # Рассчитываем процент превышения
    excess_percent = ((yandex_price - min_tender_price) / min_tender_price) * 100
    
    if excess_percent <= 5:
        # Нет превышения или до 5% - ЗЕЛЕНЫЙ
        return "00FF00"
    elif excess_percent <= 10:
        # Превышение 5-10% - ЖЕЛТЫЙ
        return "FFFF00"
    else:
        # Превышение > 10% - КРАСНЫЙ
        return "FF0000"

def debug_print_excel_rows(path: str, n: int = 50):
    """Отладка - печать первых n строк Excel"""
    try:
        df = pd.read_excel(path, header=None)
        print(f"📋 Первые {n} строк Excel: {path}")
        for i, row in df.head(n).iterrows():
            print(f"{i}: {[str(x) for x in row.tolist()]}")
        print("-" * 60)
    except Exception as e:
        print(f"❌ Ошибка чтения Excel: {e}")

def extract_products_from_excel(path: str):
    """ОРИГИНАЛЬНАЯ функция извлечения товаров из Excel"""
    all_sheets = pd.read_excel(path, header=None, sheet_name=None)
    found_df = None
    col_index = None
    start_row = None
    end_row = None
    
    print(f"📋 Анализирую Excel файл: {path}")
    
    for sheet_name, df in all_sheets.items():
        print(f"   Лист: {sheet_name}")
        for i, row in df.iterrows():
            for j, val in row.items():
                if isinstance(val, str) and 'наименование' in val.lower():
                    found_df = df
                    col_index = j
                    start_row = i + 1
                    print(f"   ✓ Найдена колонка 'Наименование' в столбце {j}, строка {i}")
                    break
            if found_df is not None:
                break
        if found_df is not None:
            break
    
    if found_df is None:
        raise ValueError("❌ Не найдена колонка 'Наименование'")
    
    for i, val in enumerate(found_df[col_index]):
        if isinstance(val, str) and 'итого' in val.lower():
            end_row = i
            break
    
    if end_row is None:
        print("⚠️ Строка 'Итого' не найдена, беру до конца данных")
        end_row = len(found_df)
    
    print(f"📊 Извлекаю товары из строк {start_row} - {end_row-1}")
    
    items = []
    for idx, text in enumerate(found_df.loc[start_row:end_row - 1, col_index]):
        if not isinstance(text, str):
            continue
        
        raw = text.strip()
        if not raw:
            continue
        
        clean_name = clean_product_name_advanced(raw)
        
        if clean_name and len(clean_name) > 3:
            items.append({
                'raw': raw,
                'name': clean_name
            })
            print(f"   {len(items)}. '{clean_name[:60]}{'...' if len(clean_name) > 60 else ''}'")
            
            if len(items) <= 5:
                if raw != clean_name:
                    print(f"      (исходно: '{raw[:40]}{'...' if len(raw) > 40 else ''}')")
    
    print(f"✅ Извлечено {len(items)} товаров")
    return pd.DataFrame(items)

def save_results_into_tender_format(original_path: str, output_path: str, df: pd.DataFrame,
                                   target_sheet_name: str = None):
    """
    ИСПРАВЛЕННАЯ функция: ищет "1 место" и сравнивает с ним
    """
    print(f"📋 Создаю тендерную таблицу с поиском '1 место'...")
    print(f"   Исходный файл: {original_path}")
    print(f"   Результат: {output_path}")
    
    try:
        shutil.copy2(original_path, output_path)
        print(f"✅ Скопирован оригинальный файл")
        
        wb = load_workbook(output_path)
        
        if target_sheet_name and target_sheet_name in wb.sheetnames:
            ws = wb[target_sheet_name]
        else:
            ws = wb.active
        
        print(f"📊 Работаю с листом: {ws.title}")
        
        name_col = None
        name_start_row = None
        
        for row_idx in range(1, 21):
            for col_idx in range(1, 11):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if isinstance(cell, MergedCell):
                    continue
                    
                cell_value = cell.value
                if cell_value and isinstance(cell_value, str) and 'наименование' in cell_value.lower():
                    name_col = col_idx
                    name_start_row = row_idx + 1
                    print(f"✅ Найдена колонка 'Наименование': колонка {get_column_letter(col_idx)}, строка {row_idx}")
                    break
            if name_col:
                break
        
        if not name_col:
            raise ValueError("❌ Не найдена колонка 'Наименование'")
        
        header_row = name_start_row - 1
        participant_columns = []
        
        for col_idx in range(name_col + 1, min(ws.max_column + 1, name_col + 15)):
            header_cell = ws.cell(row=header_row, column=col_idx)
            
            if isinstance(header_cell, MergedCell):
                continue
                
            header_value = header_cell.value
            if header_value and isinstance(header_value, str) and header_value.strip():
                has_data = False
                for check_row in range(name_start_row, min(name_start_row + 5, ws.max_row + 1)):
                    check_cell = ws.cell(row=check_row, column=col_idx)
                    if not isinstance(check_cell, MergedCell) and check_cell.value:
                        has_data = True
                        break
                
                if has_data:
                    participant_columns.append({
                        'column': col_idx,
                        'name': header_value.strip(),
                        'letter': get_column_letter(col_idx)
                    })
        
        print(f"📊 Найдено участников тендера: {len(participant_columns)}")
        for p in participant_columns:
            print(f"   - {p['name']} (колонка {p['letter']})")
        
        if participant_columns:
            last_participant_col = max([p['column'] for p in participant_columns])
            yandex_col = last_participant_col + 1
        else:
            yandex_col = name_col + 1
        
        yandex_col_letter = get_column_letter(yandex_col)
        print(f"🎯 Колонка 'Яндекс Маркет': {yandex_col_letter}")
        
        header_cell = ws.cell(row=header_row, column=yandex_col)
        
        if not isinstance(header_cell, MergedCell):
            header_cell.value = "Яндекс Маркет"
            header_cell.font = Font(bold=True)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            print(f"✅ Создан заголовок в {yandex_col_letter}{header_row}")
        else:
            print(f"⚠️ Заголовочная ячейка объединена, пропускаем")
        
        items_end_row = None
        for row_idx in range(name_start_row, min(ws.max_row + 1, name_start_row + 200)):
            name_cell = ws.cell(row=row_idx, column=name_col)
            if not isinstance(name_cell, MergedCell):
                name_value = name_cell.value
                if name_value and isinstance(name_value, str) and 'итого' in name_value.lower():
                    items_end_row = row_idx - 1
                    break
        
        if not items_end_row:
            items_end_row = name_start_row + (len(df) * 12) - 1
        
        print(f"📊 Строки с товарами: {name_start_row} - {items_end_row}")
        
        # ФИКСИРОВАННАЯ высота блока: 12 строк
        item_height = 12
        
        print(f"📏 Высота блока товара: {item_height} строк (ФИКСИРОВАННАЯ)")
        
        def safe_write_cell(row, col, value, font=None, alignment=None, fill=None):
            try:
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    return False
                
                cell.value = value
                if font:
                    cell.font = font
                if alignment:
                    cell.alignment = alignment
                if fill:
                    cell.fill = fill
                return True
            except Exception as e:
                return False
        
        filled_count = 0
        
        for idx, (_, parsed_item) in enumerate(df.iterrows()):
            base_row = name_start_row + (idx * item_height)
            
            if base_row > items_end_row:
                break
            
            price_without_nds = parsed_item.get('цена', '')
            price_with_nds = parsed_item.get('цена для юрлиц', '')
            link = parsed_item.get('ссылка', '')
            
            # Автоматический расчет НДС
            if not price_with_nds and price_without_nds:
                try:
                    clean_price = re.sub(r'[^\d,.]', '', price_without_nds)
                    clean_price = clean_price.replace(',', '.')
                    price_num = float(clean_price) if clean_price else 0
                    if price_num > 0:
                        price_with_nds_num = price_num * 1.2
                        price_with_nds = f"{price_with_nds_num:,.0f} ₽".replace(',', ' ')
                except:
                    price_with_nds = ""
            
            # ИЩЕМ КОЛОНКУ С "1 МЕСТО" (победитель тендера)
            winner_col = None
            min_price_without_nds = float('inf')
            min_price_with_nds = float('inf')
            
            # Проходим по всем участникам и ищем "1 место"
            for participant in participant_columns:
                part_col = participant['column']
                
                # Строка 1: "Ранг по цене" (base_row + 0)
                rank_row = base_row
                rank_cell = ws.cell(row=rank_row, column=part_col)
                
                if not isinstance(rank_cell, MergedCell):
                    rank_value = rank_cell.value
                    if rank_value and isinstance(rank_value, str):
                        # Ищем "1 место" или "3 место" и т.д.
                        if '1' in rank_value and 'место' in rank_value.lower():
                            winner_col = part_col
                            print(f"\n   ✅ НАЙДЕН ПОБЕДИТЕЛЬ '1 место': {participant['name']} (колонка {participant['letter']})")
                            break
            
            if winner_col is None:
                print(f"\n   ⚠️ НЕ НАЙДЕН победитель с '1 место', ищем минимум по всем")
                # Если не нашли "1 место", ищем минимум по всем участникам
                for participant in participant_columns:
                    part_col = participant['column']
                    
                    # Строка 2: Цена БЕЗ НДС (base_row + 1)
                    price_without_nds_row = base_row + 1
                    cell_without_nds = ws.cell(row=price_without_nds_row, column=part_col)
                    
                    if not isinstance(cell_without_nds, MergedCell):
                        cell_value = cell_without_nds.value
                        if cell_value and isinstance(cell_value, (int, float, str)):
                            price_val = parse_price_value(str(cell_value))
                            if price_val > 10 and price_val < 1000000:
                                if price_val < min_price_without_nds:
                                    min_price_without_nds = price_val
                    
                    # Строка 3: Цена С НДС (base_row + 2)
                    price_with_nds_row = base_row + 2
                    cell_with_nds = ws.cell(row=price_with_nds_row, column=part_col)
                    
                    if not isinstance(cell_with_nds, MergedCell):
                        cell_value = cell_with_nds.value
                        if cell_value and isinstance(cell_value, (int, float, str)):
                            price_val = parse_price_value(str(cell_value))
                            if price_val > 10 and price_val < 1000000:
                                if price_val < min_price_with_nds:
                                    min_price_with_nds = price_val
            else:
                # Читаем цены из колонки победителя
                # Строка 2: Цена БЕЗ НДС (base_row + 1)
                price_without_nds_row = base_row + 1
                cell_without_nds = ws.cell(row=price_without_nds_row, column=winner_col)
                
                if not isinstance(cell_without_nds, MergedCell):
                    cell_value = cell_without_nds.value
                    if cell_value and isinstance(cell_value, (int, float, str)):
                        price_val = parse_price_value(str(cell_value))
                        if price_val > 0:
                            min_price_without_nds = price_val
                            print(f"   Цена БЕЗ НДС победителя: {min_price_without_nds}")
                
                # Строка 3: Цена С НДС (base_row + 2)
                price_with_nds_row = base_row + 2
                cell_with_nds = ws.cell(row=price_with_nds_row, column=winner_col)
                
                if not isinstance(cell_with_nds, MergedCell):
                    cell_value = cell_with_nds.value
                    if cell_value and isinstance(cell_value, (int, float, str)):
                        price_val = parse_price_value(str(cell_value))
                        if price_val > 0:
                            min_price_with_nds = price_val
                            print(f"   Цена С НДС победителя: {min_price_with_nds}")
            
            # Если не нашли цены, используем значения по умолчанию
            if min_price_without_nds == float('inf'):
                min_price_without_nds = 0.0
            if min_price_with_nds == float('inf'):
                min_price_with_nds = 0.0
            
            # Парсим цены Яндекс Маркет
            yandex_price_without_nds_val = parse_price_value(price_without_nds)
            yandex_price_with_nds_val = parse_price_value(price_with_nds)
            
            # Определяем цвета для ячеек
            color_without_nds = get_color_by_comparison(yandex_price_without_nds_val, min_price_without_nds)
            color_with_nds = get_color_by_comparison(yandex_price_with_nds_val, min_price_with_nds)
            
            print(f"🔄 Товар {idx + 1}: {parsed_item['наименование'][:30]}...")
            print(f"   Цена БЕЗ НДС победителя: {min_price_without_nds:.2f}")
            print(f"   Цена ЯМ БЕЗ НДС: {yandex_price_without_nds_val:.2f} → #{color_without_nds}")
            print(f"   Цена С НДС победителя: {min_price_with_nds:.2f}")
            print(f"   Цена ЯМ С НДС: {yandex_price_with_nds_val:.2f} → #{color_with_nds}")
            
            if price_without_nds or price_with_nds:
                success_count = 0
                
                # Строка 2: Цена БЕЗ НДС с цветом
                if price_without_nds:
                    fill_without_nds = PatternFill(start_color=color_without_nds, 
                                                   end_color=color_without_nds, 
                                                   fill_type="solid")
                    if safe_write_cell(base_row + 2, yandex_col, price_without_nds,
                                      alignment=Alignment(horizontal='right'),
                                      fill=fill_without_nds):
                        success_count += 1
                
                # Строка 3: Цена С НДС с цветом
                if price_with_nds:
                    fill_with_nds = PatternFill(start_color=color_with_nds, 
                                               end_color=color_with_nds, 
                                               fill_type="solid")
                    if safe_write_cell(base_row + 3, yandex_col, price_with_nds,
                                      alignment=Alignment(horizontal='right'),
                                      fill=fill_with_nds):
                        success_count += 1
                
                # Строки 4-11: пусто
                for offset in range(4, 12):
                    safe_write_cell(base_row + offset, yandex_col, "",
                                   alignment=Alignment(horizontal='center'))
                
                # Строка 12: ССЫЛКА
                if link:
                    link_cell = ws.cell(row=base_row + 12, column=yandex_col)
                    if not isinstance(link_cell, MergedCell):
                        link_cell.value = "ССЫЛКА"
                        link_cell.hyperlink = link
                        link_cell.font = Font(color="0000FF", underline="single", size=9)
                        link_cell.alignment = Alignment(horizontal='center')
                        success_count += 1
                
                if success_count > 0:
                    filled_count += 1
        
        print(f"✅ Заполнено товаров: {filled_count}")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        border_count = 0
        for row_idx in range(header_row, items_end_row + 50):
            cell = ws.cell(row=row_idx, column=yandex_col)
            if not isinstance(cell, MergedCell):
                if not cell.border or cell.border == Border():
                    cell.border = thin_border
                    border_count += 1
        
        print(f"✅ Применены границы к {border_count} ячейкам")
        
        wb.save(output_path)
        print(f"💾 Тендерная таблица сохранена: {output_path}")
        
        return True
        
    except Exception as e:
        print(f"❌ Ошибка создания тендерной таблицы: {e}")
        import traceback
        traceback.print_exc()
        return False

def save_results_into_excel(original_path: str, output_path: str, df: pd.DataFrame,
                           original_sheet_name="Original", prices_sheet_name="Prices"):
    """СТАРАЯ функция сохранения (для совместимости)"""
    
    try:
        original = pd.read_excel(original_path, header=None)
    except Exception as e:
        print(f"⚠️ Ошибка чтения оригинала: {e}")
        original = pd.DataFrame()
    
    final_df = df[['наименование', 'цена', 'цена для юрлиц', 'ссылка']].copy()
    final_df.columns = ['Название', 'Цена', 'Для юрлиц', 'Ссылка']
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        if not original.empty:
            original.to_excel(writer, sheet_name=original_sheet_name, index=False, header=False)
        final_df.to_excel(writer, sheet_name=prices_sheet_name, index=False)
    
    try:
        wb = load_workbook(output_path)
        ws = wb[prices_sheet_name]
        
        if 'Ссылка' in final_df.columns:
            link_col = list(final_df.columns).index('Ссылка') + 1
            for row in range(2, ws.max_row + 1):
                url = ws.cell(row=row, column=link_col).value
                if url and isinstance(url, str) and url.strip():
                    cell = ws.cell(row=row, column=link_col)
                    cell.value = url
                    cell.hyperlink = url
                    cell.font = Font(color="0000FF", underline="single")
        
        wb.save(output_path)
        print(f"💾 Результаты сохранены: {output_path}")
    except Exception as e:
        print(f"⚠️ Ошибка добавления ссылок: {e}")

def save_cookies_pickle(driver, path: str):
    try:
        cookies = driver.get_cookies()
        with open(path, 'wb') as f:
            pickle.dump(cookies, f)
        return True
    except Exception:
        return False

def load_cookies_pickle(driver, path: str, domain_filter: str = None):
    if not os.path.exists(path):
        return False
    
    try:
        with open(path, 'rb') as f:
            cookies = pickle.load(f)
        
        for c in cookies:
            if domain_filter and 'domain' in c and domain_filter not in c['domain']:
                continue
            
            if 'expiry' in c:
                try:
                    c['expiry'] = int(c['expiry'])
                except Exception:
                    c.pop('expiry', None)
            
            try:
                driver.add_cookie(c)
            except Exception:
                continue
        
        return True
    except Exception:
        return False

def check_cookies_validity(cookies_file: str) -> dict:
    result = {
        'exists': False,
        'valid': False,
        'expired': False,
        'size': 0,
        'domains': [],
        'message': ''
    }
    
    if not os.path.exists(cookies_file):
        result['message'] = "Файл cookies не найден"
        return result
    
    result['exists'] = True
    
    try:
        with open(cookies_file, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            result['size'] = len(content)
            
            if not content:
                result['message'] = "Файл cookies пустой"
                return result
            
            import json
            cookies_data = json.loads(content)
            
            if isinstance(cookies_data, list):
                cookies = cookies_data
            elif isinstance(cookies_data, dict) and 'cookies' in cookies_data:
                cookies = cookies_data['cookies']
            else:
                result['message'] = "Неверный формат cookies"
                return result
            
            if not cookies:
                result['message'] = "Список cookies пустой"
                return result
            
            domains = set()
            expired_count = 0
            import time
            current_time = time.time()
            
            for cookie in cookies:
                if isinstance(cookie, dict) and 'domain' in cookie:
                    domains.add(cookie['domain'])
                
                if 'expirationDate' in cookie:
                    try:
                        exp_time = float(cookie['expirationDate'])
                        if exp_time < current_time:
                            expired_count += 1
                    except:
                        pass
            
            result['domains'] = list(domains)
            result['valid'] = len(cookies) > 0
            
            if expired_count > len(cookies) * 0.5:
                result['expired'] = True
                result['message'] = f"Много просроченных cookies ({expired_count}/{len(cookies)})"
            else:
                result['message'] = f"Cookies валидны ({len(cookies)} шт)"
            
    except Exception as e:
        result['message'] = f"Ошибка проверки cookies: {e}"
    
    return result
